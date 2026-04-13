from copy import copy
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_SHEET_FOR_PART = {
    "plate":     "Plates",
    "ubolt":     "Ubolt",
    "clamp":     "Clamp",
    "lug":       "Stright_Pipe_Lug",
    "elbow_lug": "Elbow_Lug",
    "bolt":      "Bolt",
    "nut":       "Nut",
    "pin":       "Pin",
    "clevis":    "Clevis",
    "rod":       "Rod",
    "shoe":      "Shoe",
    "strap":     "Strap",
    "shield":    "Shield",
    "guide":     "Guide",
    "strut":     "Strut_A",
}


def _find_markers(ws):
    """Return (start_row, end_row) for the 'Start' and 'End' markers in column 1."""
    start_row = end_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v == "Start":
            start_row = r
        elif v == "End":
            end_row = r
    if not start_row or not end_row:
        raise ValueError("Start/End markers not found in sheet")
    return start_row, end_row


def _read_headers(ws, header_row):
    """Return {header_name: col_index} from the given row."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v:
            headers[str(v)] = col
    return headers


def _find_template_row(ws, start_row, end_row):
    """
    Locate the template row between Start and End.

    Strategy (in order):
    1. Any row whose column-1 value is '!' (explicit template marker).
    2. The last non-empty row before End (fallback for existing templates).
    """
    # Pass 1: explicit '!' marker
    for r in range(start_row + 1, end_row):
        if ws.cell(r, 1).value == "!":
            return r

    # Pass 2: last row with any content before End
    for r in range(end_row - 1, start_row, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r

    raise ValueError("No template row found between Start and End")


def _read_row_values(ws, row_num):
    """Return a dict {col_index: value} for every column in the given row."""
    return {c: ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)}


def _copy_row(ws, src_row, dst_row):
    """Copy cell values and styles from src_row into dst_row."""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def _fill_ws(ws, mapped_data):
    """
    Write mapped_data rows into an already-open worksheet.
    Template row (marked '!' or last row before End) is copied for every record
    so all default values are preserved.
    """
    start_row, end_row = _find_markers(ws)
    headers = _read_headers(ws, start_row - 2)

    tmpl_row_num = _find_template_row(ws, start_row, end_row)
    tmpl_values = _read_row_values(ws, tmpl_row_num)
    logger.info(f"  Template row: {tmpl_row_num}, inserting {len(mapped_data)} record(s)")

    # Clear existing data area
    data_area_count = end_row - start_row - 1
    if data_area_count > 0:
        ws.delete_rows(start_row + 1, data_area_count)

    insert_at = start_row + 1
    for i, record in enumerate(mapped_data):
        ws.insert_rows(insert_at)

        for col, val in tmpl_values.items():
            ws.cell(insert_at, col).value = val
        ws.cell(insert_at, 1).value = None  # column 1 must be empty on data rows

        for field_name, value in record.items():
            if field_name in headers:
                ws.cell(insert_at, headers[field_name]).value = value
                logger.debug(f"  [{i+1}] {field_name} = {value}")

        logger.info(f"  Row {i+1}: PartNumber={record.get('PartNumber', '?')}")
        insert_at += 1


def fill_workbook(template_path, output_path, batches):
    """
    Write multiple batches into a single workbook — one sheet per image.

    batches: list of (mapped_data, part_type, label) tuples.
      - label is used as the sheet name (filename stem, truncated to 31 chars).

    The saved workbook keeps 'CustomInterfaces' plus one new sheet per batch.
    The original template part sheets (Plates, Ubolt, Clamp, Stright_Pipe_Lug) are removed.
    """
    wb = load_workbook(template_path)

    # Track new sheet names to guarantee uniqueness
    used_names = set(wb.sheetnames)

    new_sheets = []
    for mapped_data, part_type, label in batches:
        base_sheet = _SHEET_FOR_PART.get(part_type.lower())
        if not base_sheet:
            raise ValueError(f"Unknown part type: {part_type}")
        if base_sheet not in wb.sheetnames:
            raise ValueError(f"Template sheet '{base_sheet}' not found")

        # Build a unique sheet name from the label (≤31 chars, Excel limit)
        raw = label[:28].strip() if label else base_sheet
        sheet_name = raw
        counter = 2
        while sheet_name in used_names:
            suffix = f"_{counter}"
            sheet_name = raw[:31 - len(suffix)] + suffix
            counter += 1
        used_names.add(sheet_name)

        # Copy the template sheet and rename it
        new_ws = wb.copy_worksheet(wb[base_sheet])
        new_ws.title = sheet_name

        logger.info(f"Filling sheet '{sheet_name}' ({part_type}, {len(mapped_data)} record(s))")
        _fill_ws(new_ws, mapped_data)
        new_sheets.append(sheet_name)

    # Remove original template part sheets — keep only CustomInterfaces + new sheets.
    # Exclude any sheet that was just created (e.g. label == template sheet name).
    sheets_to_remove = set(_SHEET_FOR_PART.values()) - set(new_sheets)
    for shname in list(wb.sheetnames):
        if shname in sheets_to_remove:
            del wb[shname]

    wb.save(output_path)
    logger.info(f"Saved {output_path}  (sheets: {list(wb.sheetnames)})")


def fill_sheet(template_path, output_path, mapped_data, part_type, label=None):
    """Single-batch convenience wrapper around fill_workbook."""
    fill_workbook(template_path, output_path, [(mapped_data, part_type, label or part_type)])
